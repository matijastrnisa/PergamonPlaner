import streamlit as st
import pandas as pd
from datetime import date, timedelta, datetime
import plotly.express as px
from openpyxl import load_workbook

# ---------------------------------------------------------
# PAGE CONFIG
# ---------------------------------------------------------
st.set_page_config(page_title="Pergamon Mini-Planer", layout="wide")
st.title("🕌 Pergamon Mini-Planer – Rollen, Personen & MP-Farben")

st.markdown("""
Diese Version erlaubt dir:

- Personen zu definieren und ihnen Rollen zuzuweisen  
- Filme mit BS-Fenster zu definieren  
- Pro Film für **jede Rolle** unterschiedliche Arbeitstage einzutragen  
- `MP.xlsx` hochzuladen (interner Kalender mit Farblegende)  
- Die App wertet die **Farben** im MP aus:

  - Links oben steht eine Legende: Name + Farbfläche  
  - Überall, wo diese Farbe im Kalender vorkommt, gilt die Person als **blockiert**  
  - Nur Tage **ohne** ihre Farbe gelten als **frei**  
""")

today = date.today()

# ---------------------------------------------------------
# HELFER: MP-KALENDER PER FARBE EINLESEN
# ---------------------------------------------------------
def load_mp_availability_by_color(mp_file, personen: list[str]):
    """
    Liest MP.xlsx mit openpyxl ein und liefert:
        availability[(Person, Datum)] = "Blockiert"
    Alle anderen Tage gelten als "Frei".

    Logik:
    - Datumszeile = die Zeile mit den meisten datetime-Werten (ab Spalte 2)
    - Personen-Legende:
        in Spalte A (1) stehen Namen wie "Anna:", "Mareike:", ...
        in Spalte B (2) ist die jeweilige Farbe als Farbfeld
        -> diese Farbe ist die Person-Farbe
    - Im Kalender:
        für jede Datumsspalte und jede Zeile:
            wenn Zellenfüllung == Personenfarbe -> Person an diesem Datum blockiert
    """
    wb = load_workbook(mp_file, data_only=True)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # 1) Datumszeile finden
    date_row_idx = None
    max_dates = -1
    for r in range(1, max_row + 1):
        count = 0
        for c in range(2, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, datetime):
                count += 1
        if count > max_dates:
            max_dates = count
            date_row_idx = r

    if date_row_idx is None or max_dates <= 0:
        return {}

    # 2) Spalten -> Datum
    date_cols = {}
    for c in range(2, max_col + 1):
        v = ws.cell(row=date_row_idx, column=c).value
        if isinstance(v, datetime):
            date_cols[c] = v.date()

    # 3) Personenfarben aus der Legende (Spalte A + Farbfeld in Spalte B)
    def norm_name(s: str) -> str:
        return s.split(":")[0].strip().lower()

    person_colors = {}
    for r in range(1, 20):  # Legende ist oben irgendwo, z.B. Zeilen 1–10
        cell_name = ws.cell(row=r, column=1).value
        if not isinstance(cell_name, str):
            continue
        base = norm_name(cell_name)
        for p in personen:
            if base == p.strip().lower():
                fill = ws.cell(row=r, column=2).fill
                col = None
                if fill and fill.fgColor and fill.fgColor.type != "indexed":
                    col = fill.fgColor.rgb
                person_colors[p] = col

    # 4) Verfügbarkeit per Farbe bestimmen
    availability = {}

    # Wir gehen alle Datumsspalten und alle Zeilen durch
    for c, d in date_cols.items():
        for r in range(1, max_row + 1):
            cell = ws.cell(row=r, column=c)
            fill = cell.fill
            col = None
            if fill and fill.fgColor and fill.fgColor.type != "indexed":
                col = fill.fgColor.rgb

            if not col or col == "00000000":
                continue  # keine relevante Farbe

            # Prüfen, ob diese Farbe zu einer Person gehört
            for person, p_color in person_colors.items():
                if p_color and col == p_color:
                    availability[(person, d)] = "Blockiert"

    return availability

# ---------------------------------------------------------
# 1️⃣ PERSONEN & ROLLEN DEFINIEREN
# ---------------------------------------------------------
st.subheader("1️⃣ Personen & Rollen")

default_personen = "Anna, Mareike, Sonja, Sophia"
personen_input = st.text_input(
    "Personen (Komma-getrennt)",
    value=default_personen
)
personen = [p.strip() for p in personen_input.split(",") if p.strip()]

if not personen:
    st.warning("Bitte mindestens eine Person eintragen.")

default_roles = ["Storyboard", "Keyframes", "Animation"]
rollen_input = st.text_input(
    "Rollen (Komma-getrennt)",
    value=", ".join(default_roles)
)
rollen = [r.strip() for r in rollen_input.split(",") if r.strip()]

if not rollen:
    st.warning("Bitte mindestens eine Rolle eintragen.")

st.markdown("#### Rollen pro Person")

person_roles = {}
for person in personen:
    person_roles[person] = st.multiselect(
        f"Rollen für **{person}**",
        options=rollen,
        default=rollen,  # standard: alle können alles, du kannst abwählen
        key=f"roles_{person}"
    )

# ---------------------------------------------------------
# 2️⃣ MP-KALENDER HOCHLADEN (FARBEN)
# ---------------------------------------------------------
st.subheader("2️⃣ MP-Kalender mit Farblegende (optional)")

mp_file = st.file_uploader(
    "MP.xlsx hochladen (interner Kalender mit farbigen Balken für Personen)",
    type=["xlsx"]
)

mp_availability = None
if mp_file is not None:
    try:
        mp_availability = load_mp_availability_by_color(mp_file, personen)
        st.success("MP.xlsx geladen. Verfügbarkeit anhand der Farben erkannt.")
        st.caption(
            "Für jede Person werden alle Tage blockiert, an denen ihre Farbe im Kalender vorkommt. "
            "Nur Tage ohne ihre Farbe gelten als frei."
        )
    except Exception as e:
        st.error(f"Fehler beim Einlesen von MP.xlsx: {e}")
        mp_availability = None

# ---------------------------------------------------------
# 3️⃣ FILME DEFINIEREN
# ---------------------------------------------------------
st.subheader("3️⃣ Filme definieren")

num_films = st.number_input(
    "Wie viele Filme möchtest du testen?",
    min_value=1,
    max_value=5,
    value=2,
    step=1
)

filme = []

for i in range(num_films):
    st.markdown(f"**Film {i+1}**")
    col1, col2, col3 = st.columns(3)

    with col1:
        name = st.text_input(
            f"Name Film {i+1}",
            value=f"Film {i+1}",
            key=f"film_name_{i}"
        )
    with col2:
        bs_start = st.date_input(
            f"BS-Start {i+1}",
            value=today + timedelta(days=7),
            key=f"bs_start_{i}"
        )
    with col3:
        bs_ende = st.date_input(
            f"BS-Ende {i+1}",
            value=today + timedelta(days=37),
            key=f"bs_ende_{i}"
        )

    if bs_ende < bs_start:
        st.error(f"Film {i+1}: BS-Ende darf nicht vor BS-Start liegen.")

    st.markdown(f"_Arbeitstage je Rolle für **{name}**:_")
    role_days = {}
    if rollen:
        cols = st.columns(len(rollen))
        for j, rolle in enumerate(rollen):
            with cols[j]:
                tage = st.number_input(
                    f"{rolle}",
                    min_value=0,
                    max_value=365,
                    value=0,
                    step=1,
                    key=f"film_{i}_role_{rolle}"
                )
                role_days[rolle] = tage
    else:
        role_days = {}

    filme.append({
        "Film": name,
        "BS_Start": bs_start,
        "BS_Ende": bs_ende,
        "Role_Days": role_days
    })

# ---------------------------------------------------------
# 4️⃣ PLANUNGS-PARAMETER
# ---------------------------------------------------------
st.subheader("4️⃣ Planungs-Parameter")

max_tage_pro_tag = st.number_input(
    "Max. Einheiten pro Person und Tag (über alle Filme/Rollen)",
    min_value=1,
    max_value=3,
    value=1
)

st.markdown("_Hinweis: Es wird **nicht in der Vergangenheit** geplant (nur ab heute)._")

# ---------------------------------------------------------
# 5️⃣ PLANUNG STARTEN
# ---------------------------------------------------------
st.subheader("5️⃣ Planung ausführen")

if st.button("🚀 Planung berechnen"):
    if not personen:
        st.error("Keine Personen definiert.")
    elif not rollen:
        st.error("Keine Rollen definiert.")
    else:
        assignments = []

        for film in filme:
            film_name = film["Film"]
            start = film["BS_Start"]
            ende = film["BS_Ende"]
            role_days = film["Role_Days"]

            # gültige Tage im BS-Fenster ab heute
            tage = []
            current = start
            while current <= ende:
                if current >= today:
                    tage.append(current)
                current += timedelta(days=1)

            if not tage:
                st.warning(f"⚠️ Film „{film_name}“: keine planbaren Tage (alles in der Vergangenheit?).")
                continue

            # pro Rolle planen
            for rolle, needed_days in role_days.items():
                remaining = int(needed_days)
                if remaining <= 0:
                    continue

                # Personen, die diese Rolle können
                passende_personen = [
                    p for p in personen
                    if rolle in person_roles.get(p, [])
                ]

                if not passende_personen:
                    st.warning(
                        f"⚠️ Film „{film_name}“: keine Person hat die Rolle „{rolle}“."
                    )
                    continue

                # Greedy: über Tage iterieren
                t_index = 0
                load = {}  # (person, datum) -> belegte Einheiten

                while remaining > 0 and t_index < len(tage):
                    d = tage[t_index]
                    for person in passende_personen:
                        # MP-Verfügbarkeit prüfen (falls vorhanden)
                        if mp_availability is not None:
                            status = mp_availability.get((person, d), "Frei")
                            if status != "Frei":
                                continue  # Person an diesem Tag blockiert

                        key = (person, d)
                        used = load.get(key, 0)
                        if used < max_tage_pro_tag and remaining > 0:
                            assignments.append({
                                "Film": film_name,
                                "Rolle": rolle,
                                "Person": person,
                                "Datum": d,
                                "Anteil": 1
                            })
                            load[key] = used + 1
                            remaining -= 1
                            if remaining <= 0:
                                break
                    t_index += 1

                if remaining > 0:
                    st.warning(
                        f"⚠️ Film „{film_name}“ / Rolle „{rolle}“: {remaining} Tage konnten nicht untergebracht werden."
                    )

        if not assignments:
            st.error("Es konnten keine Zuteilungen erzeugt werden.")
        else:
            df_assign = pd.DataFrame(assignments)

            # zusätzliche Spalte für Gantt: Film + Rolle
            df_assign["Film_Rolle"] = df_assign["Film"] + " – " + df_assign["Rolle"]

            # Ergebnis-Tabelle
            st.subheader("📘 Ergebnis – Zuteilungen")
            st.dataframe(df_assign, use_container_width=True)

            # -------------------------------------------------
            # GANTT-DIAGRAMM
            # -------------------------------------------------
            st.subheader("📊 Gantt-Diagramm")

            df_gantt = df_assign.copy()
            df_gantt["Start"] = pd.to_datetime(df_gantt["Datum"])
            df_gantt["Ende"] = df_gantt["Start"] + pd.to_timedelta(1, unit="D")

            try:
                fig = px.timeline(
                    df_gantt,
                    x_start="Start",
                    x_end="Ende",
                    y="Film_Rolle",  # Film + Rolle als eigene Zeile
                    color="Person",
                    title="Pergamon Mini-Planer – Verteilung nach Film/Rolle"
                )
                fig.update_yaxes(autorange="reversed")
                st.plotly_chart(fig, use_container_width=True)
            except Exception as e:
                st.error(f"Fehler beim Erzeugen des Gantt-Diagramms: {e}")

            # -------------------------------------------------
            # EXPORT
            # -------------------------------------------------
            st.subheader("📥 Export")

            out = df_assign.copy()
            out["Datum"] = out["Datum"].astype(str)
            csv_bytes = out.to_csv(index=False).encode("utf-8")

            st.download_button(
                "Zuteilungen als CSV herunterladen",
                data=csv_bytes,
                file_name="Pergamon_MultiRole_Mit_MP_Farben_Zuteilungen.csv",
                mime="text/csv"
            )
